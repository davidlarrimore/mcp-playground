import logging
import os
import sys
from pathlib import Path
from typing import List, Optional
from fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment

# Add parent directory to path for shared module
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from shared.download_urls import generate_signed_url

LOG_LEVEL = os.getenv("LOG_LEVEL", "INFO").upper()
logging.basicConfig(level=LOG_LEVEL, format="%(asctime)s [%(levelname)s] %(name)s - %(message)s")
logger = logging.getLogger("excel-mcp")

WORKSPACE = Path(os.getenv("WORKSPACE", "/workspace")).resolve()

mcp = FastMCP("excel")

@mcp.tool
async def save_uploaded_file(
    filename: str,
    content_base64: str
) -> dict:
    """
    Save an uploaded file (provided as base64) to the workspace.

    This is a convenience tool for when users upload files via chat.
    After saving, the file will be accessible to all workspace tools.

    IMPORTANT: Preserve the original file extension! If user uploads "data.xlsx",
    save it as "data.xlsx" (or similar .xlsx name), NOT as "data.txt".

    Args:
        filename: Name to save the file as - MUST keep original file extension
        content_base64: Base64-encoded file content from uploaded file
    """
    import base64

    try:
        file_data = base64.b64decode(content_base64)
        output_path = WORKSPACE / filename
        output_path.write_bytes(file_data)

        logger.info(f"Saved uploaded file: {output_path}")

        return {
            "path": str(output_path.relative_to(WORKSPACE)),
            "absolute_path": str(output_path),
            "size": output_path.stat().st_size,
            "message": f"File saved successfully as '{filename}'"
        }
    except Exception as e:
        error_msg = f"Error saving file: {str(e)}"
        logger.error(error_msg)
        return {"error": error_msg}

@mcp.tool
async def create_excel_workbook(
    sheets: List[dict],
    output_filename: str,
    add_charts: bool = True
) -> dict:
    """
    Create an Excel workbook with multiple sheets.

    Args:
        sheets: List of sheets with structure:
            [{
                "name": "Sheet1",
                "data": {
                    "columns": ["Region", "Revenue", "Orders"],
                    "rows": [
                        {"Region": "East", "Revenue": 50000, "Orders": 120},
                        {"Region": "West", "Revenue": 75000, "Orders": 180}
                    ]
                },
                "chart": {
                    "type": "bar",
                    "title": "Revenue by Region"
                }
            }]
        output_filename: Name of output file
        add_charts: Whether to add charts
    """
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets:
        ws = wb.create_sheet(title=sheet_def["name"])
        data = sheet_def["data"]

        # Write headers
        for col_idx, col_name in enumerate(data["columns"], 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row_data in enumerate(data["rows"], 2):
            for col_idx, col_name in enumerate(data["columns"], 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(col_name))

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = max_length + 2

        # Add chart if requested
        if add_charts and "chart" in sheet_def:
            chart = BarChart() if sheet_def["chart"].get("type") == "bar" else LineChart()
            chart.title = sheet_def["chart"].get("title", "Chart")

            data_ref = Reference(ws, min_col=2, min_row=1, max_row=len(data["rows"]) + 1)
            cats = Reference(ws, min_col=1, min_row=2, max_row=len(data["rows"]) + 1)

            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            ws.add_chart(chart, "E5")

    output_path = WORKSPACE / output_filename
    wb.save(output_path)

    logger.info(f"Created Excel workbook: {output_path}")

    # Generate signed download URL
    try:
        download_info = generate_signed_url(str(output_path))
        download_url = download_info["url"]
        expires_hours = download_info["expires_in"] // 3600
    except Exception as e:
        logger.error(f"Error generating download URL: {e}")
        download_info = None
        download_url = None
        expires_hours = 0

    result = {
        "path": str(output_path.relative_to(WORKSPACE)),
        "absolute_path": str(output_path),
        "size": output_path.stat().st_size,
        "sheets": len(sheets),
        "message": f"Excel workbook created successfully with {len(sheets)} sheet(s)."
    }

    if download_info:
        result["download_url"] = download_url
        result["download_expires_at"] = download_info["expires_at"]
        result["download_expires_in"] = download_info["expires_in"]
        result["message"] += f"\n\n📥 Download your file:\n{download_url}\n\n⏰ Link expires in {expires_hours} hours"

    return result

if __name__ == "__main__":
    mcp.run()
